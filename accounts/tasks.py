# django_ma/accounts/tasks.py
from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from celery import shared_task
from django.conf import settings
from django.core.cache import cache
from django.db import transaction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from .constants import (
    CACHE_ERROR_PREFIX,
    CACHE_PROGRESS_PREFIX,
    CACHE_RESULT_PATH_PREFIX,
    CACHE_STATUS_PREFIX,
    CACHE_TIMEOUT_SECONDS,
    cache_key,
)
from .models import CustomUser
from .services.users_excel_import import (
    REQUIRED_COLS,
    build_defaults_from_row,
    pick_worksheet_by_required_cols,
)

logger = logging.getLogger(__name__)


# =============================================================================
# 0) Ï†ïÏ±Ö/ÏÉÅÏàò
# =============================================================================

# ‚úÖ Í¥ÄÎ¶¨Ïûê Î≥¥Ìò∏(Í∂åÏû•): Í∏∞Ï°¥ Ïù¥ Îì±Í∏âÏùÄ ÏóëÏÖÄÎ°ú grade Í∞ïÎì±/Í∂åÌïú ÌïÑÎìú ÎçÆÏñ¥Ïì∞Í∏∞ Î∞©ÏßÄ
PROTECTED_GRADES = {"superuser", "head", "leader"}

# ‚úÖ Î≥¥Ìò∏ ÌïÑÎìú: ÌèâÏÉÅÏãú ÎπàÍ∞íÏúºÎ°ú ÎçÆÏñ¥Ïì∞ÏßÄ ÏïäÏùå(Îã®, Ïû¨ÏßÅ‚ÜíÌá¥ÏÇ¨ Ï†ÑÌôò Ïãú Ï¥àÍ∏∞Ìôî ÌóàÏö©)
PROTECTED_FIELDS = {"position", "team_a", "team_b", "team_c"}

# Í≤∞Í≥º Î¶¨Ìè¨Ìä∏ ÏóëÏÖÄ ÏãúÌä∏Î™Ö
RESULT_SHEET_NAME = "UploadResult"

# ÏßÑÌñâÎ•† ÌëúÏãúÎ•º ÏúÑÌïú ÏµúÏÜå/ÏµúÎåÄ Î≥¥Ï†ï
PERCENT_MIN = 0
PERCENT_MAX = 100


# =============================================================================
# 1) Cache helpers (keys Îã®ÏùºÌôî)
# =============================================================================

@dataclass(frozen=True)
class UploadCacheKeys:
    percent: str
    status: str
    error: str
    result_path: str


def _keys(task_id: str) -> UploadCacheKeys:
    return UploadCacheKeys(
        percent=cache_key(CACHE_PROGRESS_PREFIX, task_id),
        status=cache_key(CACHE_STATUS_PREFIX, task_id),
        error=cache_key(CACHE_ERROR_PREFIX, task_id),
        result_path=cache_key(CACHE_RESULT_PATH_PREFIX, task_id),
    )


def _cache_init(task_id: str) -> UploadCacheKeys:
    k = _keys(task_id)
    cache.set(k.status, "RUNNING", timeout=CACHE_TIMEOUT_SECONDS)
    cache.set(k.percent, 0, timeout=CACHE_TIMEOUT_SECONDS)
    cache.delete(k.error)
    cache.delete(k.result_path)
    return k


def _cache_set_percent(k: UploadCacheKeys, percent: int) -> None:
    p = max(PERCENT_MIN, min(PERCENT_MAX, int(percent)))
    cache.set(k.percent, p, timeout=CACHE_TIMEOUT_SECONDS)


def _cache_fail(k: UploadCacheKeys, err: Exception) -> None:
    cache.set(k.status, "FAILURE", timeout=CACHE_TIMEOUT_SECONDS)
    cache.set(k.error, str(err), timeout=CACHE_TIMEOUT_SECONDS)


def _cache_success(k: UploadCacheKeys, result_path: str) -> None:
    _cache_set_percent(k, 100)
    cache.set(k.status, "SUCCESS", timeout=CACHE_TIMEOUT_SECONDS)
    cache.set(k.result_path, result_path, timeout=CACHE_TIMEOUT_SECONDS)


# =============================================================================
# 2) Result dir helper
# =============================================================================

def _get_result_dir() -> Path:
    media_root = Path(getattr(settings, "MEDIA_ROOT", "media"))
    default_dir = media_root / "upload_results"
    result_dir = Path(getattr(settings, "UPLOAD_RESULT_DIR", default_dir))
    result_dir.mkdir(parents=True, exist_ok=True)
    return result_dir


# =============================================================================
# 3) Í≤∞Í≥º Î¶¨Ìè¨Ìä∏ ÏóëÏÖÄ ÏÉùÏÑ±
# =============================================================================

def _make_result_wb(
    results: List[List[Any]],
    total: int,
    new_cnt: int,
    upd_cnt: int,
    skip_cnt: int,
    err_cnt: int,
    picked_sheet: str,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = RESULT_SHEET_NAME

    ws.append(["Row", "ÏÇ¨ÏõêÎ≤àÌò∏", "ÏÑ±Î™Ö", "Î∂ÄÎ¨∏", "Î∂ÄÏÑú", "ÏßÄÏ†ê", "Í∂åÌïú(grade)", "ÏÉÅÌÉú", "Result"])

    fill_new = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_update = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    fill_skip = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in results:
        ws.append(row)
        r = ws.max_row
        t = str(row[-1] or "")
        cell = ws[f"I{r}"]
        if "üü¢" in t:
            cell.fill = fill_new
        elif "‚úÖ" in t:
            cell.fill = fill_update
        elif "‚ö†Ô∏è" in t:
            cell.fill = fill_skip
        elif "‚ùå" in t:
            cell.fill = fill_error

    ws.append([])
    ws.append(["ÏÑ†ÌÉùÎêú ÏãúÌä∏", picked_sheet])
    ws.append(["Ï¥ù Îç∞Ïù¥ÌÑ∞(Ìñâ)", total])
    ws.append(["Ïã†Í∑ú Ï∂îÍ∞Ä", new_cnt])
    ws.append(["ÏóÖÎç∞Ïù¥Ìä∏", upd_cnt])
    ws.append(["Ïä§ÌÇµ", skip_cnt])
    ws.append(["Ïò§Î•ò", err_cnt])

    return wb


def _save_result_workbook(task_id: str, result_wb: Workbook) -> str:
    result_dir = _get_result_dir()
    filename = f"upload_result_{task_id}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = result_dir / filename
    result_wb.save(str(path))
    return str(path)


# =============================================================================
# 4) Celery Task: ÏòÅÏóÖÍ∞ÄÏ°±ÏßÅÏõêÏ°∞Ìöå ÏóÖÎ°úÎìú/ÏóÖÎç∞Ïù¥Ìä∏
# =============================================================================

@shared_task(bind=True)
def process_users_excel_task(self, task_id: str, file_path: str, batch_size: int = 500) -> dict:
    """
    ‚úÖ 'ÏòÅÏóÖÍ∞ÄÏ°±ÏßÅÏõêÏ°∞Ìöå' ÏóëÏÖÄ ÏóÖÎ°úÎìú/ÏóÖÎç∞Ïù¥Ìä∏ (SSOT)

    - ÏãúÌä∏Î™Ö Î¨¥Í¥Ä: REQUIRED_COLS Ìè¨Ìï® ÏãúÌä∏ ÏûêÎèô ÌÉêÏÉâ
    - division(Ï¥ùÍ¥Ñ): Îπà Î¨∏ÏûêÏó¥ Ï†ÄÏû•
    - is_staff: Ï†ÑÏ≤¥ False / is_superuser: Í∏∞Î≥∏ False
    - is_active: grade != inactive
    - Í¥ÄÎ¶¨Ïûê Î≥¥Ìò∏(Í∂åÏû•): Í∏∞Ï°¥ superuser/head/leaderÎäî grade/status/is_staff/is_superuser/is_active ÎçÆÏñ¥Ïì∞Í∏∞ Í∏àÏßÄ
    - Î≥¥Ìò∏ÌïÑÎìú(PROTECTED_FIELDS): ÎπàÍ∞íÏúºÎ°ú ÎçÆÏñ¥Ïì∞ÏßÄ ÏïäÏùå (Îã®, Ïû¨ÏßÅ‚ÜíÌá¥ÏÇ¨ Ï†ÑÌôò Ïãú Ï¥àÍ∏∞Ìôî ÌóàÏö©)
    - ÏßÑÌñâÎ•†/ÏÉÅÌÉú/Ïò§Î•ò/Í≤∞Í≥ºÍ≤ΩÎ°ú: cache Í∏∞Î°ù
    - Î∞∞Ïπò Ï≤òÎ¶¨: batch_size Îã®ÏúÑ transaction
    - Í≤∞Í≥º Î¶¨Ìè¨Ìä∏ ÏóëÏÖÄ Ï†ÄÏû•
    """
    k = _cache_init(task_id)
    logger.warning("[TASK START] tid=%s file=%s batch=%s", task_id, file_path, batch_size)

    wb = None
    try:
        # ---------------------------------------------------------------------
        # 1) Workbook open + ÏóÖÎ°úÎìú ÏãúÌä∏ ÏûêÎèô ÏÑ†ÌÉù
        # ---------------------------------------------------------------------
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_name, ws, headers = pick_worksheet_by_required_cols(wb)

        if ws.sheet_state in ("hidden", "veryHidden"):
            raise ValueError("ÏóÖÎ°úÎìú ÏãúÌä∏Í∞Ä Ïà®ÍπÄ ÏÉÅÌÉúÏûÖÎãàÎã§. Ïà®ÍπÄ Ìï¥Ï†ú ÌõÑ ÏóÖÎ°úÎìúÌïòÏÑ∏Ïöî.")

        header_set = set(headers)
        missing = [c for c in REQUIRED_COLS if c not in header_set]
        if missing:
            raise ValueError(f"ÌïÑÏàò Ïª¨Îüº ÎàÑÎùΩ: {', '.join(missing)} (ÏãúÌä∏: {sheet_name})")

        total = max(int(ws.max_row) - 1, 0)  # Ìó§Îçî Ï†úÏô∏

        # ---------------------------------------------------------------------
        # 2) ÏÇ¨ÏõêÎ≤àÌò∏ ÏÑ† ÏàòÏßë ‚Üí Í∏∞Ï°¥ ÏÇ¨Ïö©Ïûê Îì±Í∏â Ï°∞Ìöå(Í¥ÄÎ¶¨Ïûê Î≥¥Ìò∏ ÌåêÎã®)
        #    (read_only iterator 1Ìöå ÏÜåÎ™® Î∞©ÏßÄÎ•º ÏúÑÌï¥: ÏÑ† ÏàòÏßë ÌõÑ workbook Ïû¨Ïò§Ìîà)
        # ---------------------------------------------------------------------
        ids: List[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            emp_id, _name, _defaults = build_defaults_from_row(headers, row)
            if emp_id:
                ids.append(emp_id)

        existing_grade_map = dict(
            CustomUser.objects.filter(id__in=ids).values_list("id", "grade")
        )

        try:
            wb.close()
        except Exception:
            pass

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_name, ws, headers = pick_worksheet_by_required_cols(wb)

        # ---------------------------------------------------------------------
        # 3) Î∞∞Ïπò Ï≤òÎ¶¨ Ï§ÄÎπÑ
        # ---------------------------------------------------------------------
        results: List[List[Any]] = []
        created = updated = skipped = err_cnt = 0
        processed = 0

        buffer_rows: List[Tuple[Any, ...]] = []
        current_excel_row_num = 2  # ÏóëÏÖÄ Ïã§Ï†ú ÌñâÎ≤àÌò∏(Ìó§Îçî Îã§Ïùå)

        def set_percent_from_processed() -> None:
            if total <= 0:
                _cache_set_percent(k, 100)
                return
            p = int((processed / total) * 100)
            _cache_set_percent(k, p)

        @transaction.atomic
        def flush_chunk(rows_chunk: List[Tuple[Any, ...]], start_row_num: int) -> None:
            nonlocal created, updated, skipped, err_cnt, processed, results, existing_grade_map

            # ÏÑ±Îä•: chunk ÎÇ¥ Í∏∞Ï°¥ user Ìïú Î≤àÏóê ÎØ∏Î¶¨ Î°úÎìú
            chunk_ids: List[str] = []
            built: List[Tuple[int, str, str, Dict[str, Any]]] = []  # (excel_row_num, emp_id, name, defaults)

            for offset, row in enumerate(rows_chunk):
                excel_row_num = start_row_num + offset
                emp_id, name, defaults = build_defaults_from_row(headers, row)

                if not emp_id:
                    skipped += 1
                    results.append([excel_row_num, "", name, "", "", "", "", "", "‚ö†Ô∏è ÏÇ¨ÏõêÎ≤àÌò∏ ÎàÑÎùΩ(Ïä§ÌÇµ)"])
                    processed += 1
                    continue

                built.append((excel_row_num, emp_id, name, defaults))
                chunk_ids.append(emp_id)

            users_by_id: Dict[str, CustomUser] = {
                u.id: u for u in CustomUser.objects.filter(id__in=chunk_ids)
            }

            for excel_row_num, emp_id, name, defaults in built:
                try:
                    user = users_by_id.get(emp_id)
                    channel = defaults.get("channel", "")
                    part = defaults.get("part", "")
                    branch = defaults.get("branch", "")
                    grade = defaults.get("grade", "")
                    status = defaults.get("status", "")
                    quit_ = defaults.get("quit")

                    # ---------------------------------------------------------
                    # Update path
                    # ---------------------------------------------------------
                    if user:
                        is_protected = user.grade in PROTECTED_GRADES
                        quit_newly_added = (user.quit is None and quit_ is not None)

                        # 1) Î≥¥Ìò∏Îì±Í∏â + Ìá¥ÏÇ¨Ïùº Ïã†Í∑ú ÏÉùÏÑ± ÏïÑÎãò ‚Üí Î≥ÄÍ≤Ω Ï∞®Îã®
                        if is_protected and not quit_newly_added:
                            skipped += 1
                            results.append([
                                excel_row_num,
                                emp_id,
                                name,
                                channel,
                                part,
                                branch,
                                getattr(user, "grade", ""),
                                getattr(user, "status", ""),
                                "‚ö†Ô∏è Î≥¥Ìò∏Îì±Í∏â(superuser/head/leader) - Ìá¥ÏÇ¨Ïùº Ïã†Í∑ú ÏóÜÏùå(Î≥ÄÍ≤Ω Ï∞®Îã®)",
                            ])
                            processed += 1
                            continue

                        # 2) Î≥¥Ìò∏Îì±Í∏â + Ìá¥ÏÇ¨Ïùº Ïã†Í∑ú ÏÉùÏÑ± ‚Üí resign/inactive Í∞ïÏ†ú Ï†ÑÌôò
                        if is_protected and quit_newly_added:
                            forced_grade = "inactive" if ((not name) or ("*" in name)) else "resign"
                            forced_status = "Ïû¨ÏßÅ" if forced_grade == "basic" else "Ìá¥ÏÇ¨"
                            defaults["grade"] = forced_grade
                            defaults["status"] = forced_status
                            defaults["is_active"] = (forced_grade != "inactive")
                            defaults["is_staff"] = False
                            defaults["is_superuser"] = False

                        update_fields: List[str] = []

                        for key, value in defaults.items():
                            # Î≥¥Ìò∏ ÌïÑÎìú: ÌèâÏÜåÏóî ÎπàÍ∞í ÎçÆÏñ¥Ïì∞Í∏∞ Í∏àÏßÄ, Îã® quit Ïã†Í∑ú ÏÉùÏÑ±Ïù¥Î©¥ Ï¥àÍ∏∞Ìôî ÌóàÏö©
                            if key in PROTECTED_FIELDS:
                                if value:
                                    setattr(user, key, value)
                                    update_fields.append(key)
                                else:
                                    if quit_newly_added:
                                        setattr(user, key, "")
                                        update_fields.append(key)
                                continue

                            # ÏùºÎ∞ò ÌïÑÎìú: None/ÎπàÎ¨∏ÏûêÏó¥ÏùÄ ÎçÆÏñ¥Ïì∞Í∏∞ ÌïòÏßÄ ÏïäÏùå(Îç∞Ïù¥ÌÑ∞ ÏÜåÏã§ Î∞©ÏßÄ)
                            if value is None:
                                continue
                            if isinstance(value, str) and value == "":
                                continue

                            setattr(user, key, value)
                            update_fields.append(key)

                        if update_fields:
                            # Ï§ëÎ≥µ Ï†úÍ±∞
                            user.save(update_fields=list(dict.fromkeys(update_fields)))

                        existing_grade_map[emp_id] = user.grade

                        updated += 1
                        results.append([
                            excel_row_num,
                            emp_id,
                            name,
                            channel,
                            part,
                            branch,
                            getattr(user, "grade", ""),
                            getattr(user, "status", ""),
                            "‚úÖ Í∏∞Ï°¥ ÏóÖÎç∞Ïù¥Ìä∏",
                        ])

                    # ---------------------------------------------------------
                    # Create path
                    # ---------------------------------------------------------
                    else:
                        CustomUser.objects.create_user(
                            id=emp_id,
                            password=emp_id,  # Ï¥àÍ∏∞ ÎπÑÎ∞ÄÎ≤àÌò∏ = ÏÇ¨ÏõêÎ≤àÌò∏
                            **defaults,
                        )
                        existing_grade_map[emp_id] = defaults.get("grade", "basic")

                        created += 1
                        results.append([
                            excel_row_num,
                            emp_id,
                            name,
                            channel,
                            part,
                            branch,
                            defaults.get("grade", ""),
                            defaults.get("status", ""),
                            "üü¢ Ïã†Í∑ú Îì±Î°ù",
                        ])

                except Exception as e:
                    err_cnt += 1
                    results.append([
                        excel_row_num,
                        emp_id,
                        name,
                        defaults.get("channel", ""),
                        defaults.get("part", ""),
                        defaults.get("branch", ""),
                        defaults.get("grade", ""),
                        defaults.get("status", ""),
                        f"‚ùå Ïò§Î•ò: {e}",
                    ])

                processed += 1

            set_percent_from_processed()

        # ---------------------------------------------------------------------
        # 4) batch loop
        # ---------------------------------------------------------------------
        for row in ws.iter_rows(min_row=2, values_only=True):
            buffer_rows.append(row)
            if len(buffer_rows) >= batch_size:
                flush_chunk(buffer_rows, start_row_num=current_excel_row_num)
                current_excel_row_num += len(buffer_rows)
                buffer_rows = []

        if buffer_rows:
            flush_chunk(buffer_rows, start_row_num=current_excel_row_num)

        # ---------------------------------------------------------------------
        # 5) Í≤∞Í≥º Î¶¨Ìè¨Ìä∏ ÏÉùÏÑ± + Ï†ÄÏû•
        # ---------------------------------------------------------------------
        result_wb = _make_result_wb(
            results=results,
            total=total,
            new_cnt=created,
            upd_cnt=updated,
            skip_cnt=skipped,
            err_cnt=err_cnt,
            picked_sheet=sheet_name,
        )
        result_path = _save_result_workbook(task_id, result_wb)

        # ---------------------------------------------------------------------
        # 6) cache finalize (SUCCESS)
        # ---------------------------------------------------------------------
        _cache_success(k, result_path)

        logger.warning(
            "[TASK DONE] tid=%s status=SUCCESS sheet=%s total=%s created=%s updated=%s skipped=%s errors=%s",
            task_id, sheet_name, total, created, updated, skipped, err_cnt
        )

        return {
            "status": "SUCCESS",
            "result_path": result_path,
            "sheet": sheet_name,
            "total": total,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": err_cnt,
        }

    except Exception as e:
        logger.exception("[TASK FAIL] tid=%s file=%s", task_id, file_path)
        _cache_fail(k, e)
        raise

    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass
